# main.py
import csv
import os
import re
from datetime import datetime
from collections import defaultdict
from grafo import construir_grafo, colorir_grafo_balanceado

# ========= Leitura dos CSVs =========

def carregar_disciplinas(caminho):
    """
    Lê 'disciplinas.csv' aceitando colunas: nome, prof (opcional), semestre (opcional).
    - 'prof' pode vir com separadores | , ; /
    - gera 'prof' normalizado com '|'
    """
    with open(caminho, newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.DictReader(csvfile)
        out = []
        for r in reader:
            nome = (r.get("nome") or r.get("disciplina") or "").strip()
            if not nome:
                continue

            # Coleta professores de colunas que comecem com 'prof' ou nomes comuns
            pedacos = []
            for k, v in r.items():
                if not k:
                    continue
                lk = k.strip().lower()
                if lk in {"prof", "professor", "professores", "docente", "docentes"} or lk.startswith("prof"):
                    if v:
                        pedacos.append(str(v))

            # Normaliza professores
            toks = []
            for p in pedacos if pedacos else [r.get("prof", "")]:
                toks += [t.strip() for t in re.split(r"[|,;/]+", str(p)) if t.strip()]
            uniq = list(dict.fromkeys(toks))  # preserva ordem e remove duplicatas
            prof_norm = "|".join(uniq)        # para o grafo
            prof_display = ", ".join(uniq)    # para exibição

            # Semestre (opcional)
            sem = ""
            for chave in ("semestre", "sem", "periodo", "período"):
                if chave in r and r[chave]:
                    sem = str(r[chave]).strip()
                    break

            out.append({
                "nome": nome,
                "prof": prof_norm,
                "prof_display": prof_display,
                "semestre": sem,
            })
        return out

def carregar_pares(caminho):
    pares = []
    if os.path.exists(caminho):
        with open(caminho, newline='', encoding='utf-8-sig') as csvfile:
            for r in csv.DictReader(csvfile):
                a = (r.get("disciplina1") or "").strip()
                b = (r.get("disciplina2") or "").strip()
                if a and b:
                    pares.append((a, b))
    return pares

def carregar_fixos(caminho):
    fixos = {}
    if os.path.exists(caminho):
        with open(caminho, newline='', encoding='utf-8-sig') as csvfile:
            for r in csv.DictReader(csvfile):
                disc = (r.get("disciplina") or "").strip()
                bloco_txt = str(r.get("bloco", "")).strip()
                if not disc or bloco_txt == "":
                    continue
                try:
                    fixos[disc] = int(bloco_txt)
                except ValueError:
                    print(f"[AVISO] Bloco inválido para '{disc}': '{bloco_txt}' (ignorado).")
    return fixos

def carregar_dia_fixo(caminho):
    """
    Aceita: seg/segunda, ter/terça/terca, qua/quarta, qui/quinta, sex/sexta,
            sab/sábado/sabado, dom/domingo
    """
    dia_por_disc = {}
    mapa = {
        'segunda':'segunda','seg':'segunda',
        'terca':'terca','terça':'terca','ter':'terca',
        'quarta':'quarta','qua':'quarta',
        'quinta':'quinta','qui':'quinta',
        'sexta':'sexta','sex':'sexta',
        'sabado':'sabado','sábado':'sabado','sab':'sabado',
        'domingo':'domingo','dom':'domingo'
    }
    if os.path.exists(caminho):
        with open(caminho, newline='', encoding='utf-8-sig') as csvfile:
            for r in csv.DictReader(csvfile):
                d = (r.get("disciplina") or "").strip()
                dia = (r.get("dia") or "").strip().lower()
                if d and dia in mapa:
                    dia_por_disc[d] = mapa[dia]
                elif d:
                    print(f"[AVISO] Dia '{dia}' inválido para {d} (ignorado).")
    return dia_por_disc

# ========= Calendário (dias × blocos/dia) =========

DIAS_ABRV = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']
MAPA_DOM  = {'seg':'segunda','ter':'terca','qua':'quarta','qui':'quinta','sex':'sexta','sáb':'sabado','sab':'sabado','dom':'domingo'}

def _gerar_horas(qtd, base_h=8, passo=2):
    # ['08:00','10:00',...] começando às 08:00 e pulando de 2h
    return [f"{(base_h + i*passo):02d}:00" for i in range(qtd)]

def montar_horarios(dias_semana:int, blocos_por_dia:int):
    if not (1 <= dias_semana <= 7):
        raise ValueError("dias_semana deve estar entre 1 e 7.")
    if blocos_por_dia <= 0:
        raise ValueError("blocos_por_dia deve ser > 0.")
    dias  = DIAS_ABRV[:dias_semana]
    horas = _gerar_horas(blocos_por_dia, base_h=8, passo=2)
    horarios, idx = {}, 0
    for d in dias:
        for h in horas:
            horarios[idx] = f"{d} {h}"
            idx += 1
    return horarios  # {bloco:int -> "Seg 08:00"}

def indice_blocos_por_dia(horarios):
    """
    Índice: dia_normalizado -> {blocos} a partir de labels "Seg 08:00", etc.
    """
    idx = {}
    for bloco, label in horarios.items():
        dia_sigla = str(label).split()[0].lower()
        if dia_sigla in MAPA_DOM:
            dia_norm = MAPA_DOM[dia_sigla]
            idx.setdefault(dia_norm, set()).add(bloco)
    return idx

# ========= Exportação: Excel (matriz/lista) + CSV =========

def salvar_grade_excel_csv(cores, horarios, nome_exibicao, caminho_base="out/alocacaohorario"):
    """
    Gera:
      - caminho_base.xlsx  (abas: 'Grade (Matriz)' e 'Grade (Lista)')
      - caminho_base.csv   (lista Dia/Hora/Disciplina)

    cores: dict {disciplina -> bloco}
    horarios: dict {bloco:int -> 'Seg 08:00', ...}
    nome_exibicao: dict {disciplina -> 'Disciplina / Professor(es)'}
    """
    try:
        import pandas as pd
    except Exception:
        print("[AVISO] Para exportar Excel/CSV instale: python -m pip install pandas openpyxl")
        return

    os.makedirs("out", exist_ok=True)

    # 1) Monta lista (Dia, Hora, Disciplina exibida)
    data = []
    for disc, bloco in cores.items():
        label = horarios.get(bloco, "")
        try:
            dia_sigla, hhmm = label.split()
        except ValueError:
            continue
        exib = nome_exibicao.get(disc, disc)  # "Disciplina / Prof(s)"
        data.append((dia_sigla, hhmm, exib))

    if not data:
        print("[AVISO] Nada para exportar (data vazia).")
        return

    # Ordem canônica de dias e horas presentes
    dias_ord_padrao = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    dias_presentes = [d for d in dias_ord_padrao if any(d == x[0] for x in data)]
    horas_presentes = sorted({h for _, h, _ in data})

    # 2) DataFrame lista
    list_df = pd.DataFrame(data, columns=["Dia", "Hora", "Disciplina / Professor(es)"])
    list_df["Dia"]  = pd.Categorical(list_df["Dia"],  categories=dias_presentes, ordered=True)
    list_df["Hora"] = pd.Categorical(list_df["Hora"], categories=horas_presentes, ordered=True)
    list_df = list_df.sort_values(["Dia", "Hora", "Disciplina / Professor(es)"]).reset_index(drop=True)

    # 3) DataFrame matriz (linhas=dia, colunas=hora; célula = itens com \n)
    matriz_dict = {d: {h: [] for h in horas_presentes} for d in dias_presentes}
    for d, h, exib in data:
        if d in matriz_dict and h in matriz_dict[d]:
            matriz_dict[d][h].append(exib)

    matriz_df = pd.DataFrame(
        {h: {d: "\n".join(matriz_dict[d][h]) if matriz_dict[d][h] else "" for d in dias_presentes}
         for h in horas_presentes}
    ).reindex(dias_presentes)
    matriz_df.index.name = "Dia"

    # Caminhos absolutos (só para imprimir bonitinho)
    base_abs = os.path.abspath(caminho_base)
    xlsx_path = f"{base_abs}.xlsx"
    csv_path  = f"{base_abs}.csv"

    # 4) Salvar Excel (com formatação básica) e CSV
    try:
        from openpyxl.styles import Alignment, Font
        openpyxl_ok = True
    except Exception:
        openpyxl_ok = False

    def _write_excel(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Aba 1: Matriz
            matriz_df.to_excel(writer, sheet_name="Grade (Matriz)")
            ws = writer.sheets["Grade (Matriz)"]
            if openpyxl_ok:
                ws.column_dimensions["A"].width = 14  # "Dia"
                for i in range(len(horas_presentes)):
                    col_letter = chr(ord("B") + i)
                    ws.column_dimensions[col_letter].width = 40
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Aba 2: Lista
            list_df.to_excel(writer, sheet_name="Grade (Lista)", index=False)
            ws2 = writer.sheets["Grade (Lista)"]
            if openpyxl_ok:
                ws2.column_dimensions["A"].width = 10  # Dia
                ws2.column_dimensions["B"].width = 8   # Hora
                ws2.column_dimensions["C"].width = 60  # Disciplina / Professor(es)
                for cell in ws2[1]:
                    cell.font = Font(bold=True)

    try:
        _write_excel(xlsx_path)
        print(f"\nExcel salvo em: {xlsx_path}")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base_abs}_{ts}.xlsx"
        _write_excel(alt)
        print(f"\n[AVISO] O Excel estava aberto. Salvei com outro nome: {alt}")

    list_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print("CSV salvo em:", csv_path)

# ========= Execução =========

def main():
    dados_dir = "dados"

    # 1) Quantos dias e quantos blocos por dia
    try:
        dias_semana = int(input("Quantos dias na semana (1–7)? ").strip() or "5")
        blocos_dia  = int(input("Quantos blocos por dia? ").strip() or "4")
    except ValueError:
        print("[AVISO] Entrada inválida. Usando padrão 5 dias × 4 blocos.")
        dias_semana, blocos_dia = 5, 4

    horarios = montar_horarios(dias_semana, blocos_dia)
    num_blocos = len(horarios)

    # 2) Entradas
    disciplinas = carregar_disciplinas(os.path.join(dados_dir, "disciplinas.csv"))

    # Grafo com conflito por professor + semestre
    G = construir_grafo(
        disciplinas,
        conflito_por_prof=True,
        conflito_por_semestre=True
    )

    # Mapa de professores por disciplina (para exibição)
    prof_display = {d["nome"]: d.get("prof_display", "") for d in disciplinas}
    nome_exibicao = {disc: (f"{disc} / {prof_display[disc]}" if prof_display.get(disc) else disc) for disc in G.nodes()}

    # Restrições "não podem coincidir" (opcional; pode remover se quiser só por semestre/prof)
    restricoes = carregar_pares(os.path.join(dados_dir, "restricoes.csv"))
    for a, b in restricoes:
        if a in G and b in G:
            G.add_edge(a, b)
        else:
            print(f"[AVISO] Restrição ignorada (nó não existe no grafo): ({a}, {b})")

    # Grupos "mesmo bloco" (opcionais)
    pares_mesmo_a = carregar_pares(os.path.join(dados_dir, "mesmo_bloco.csv"))
    pares_mesmo_b = carregar_pares(os.path.join(dados_dir, "mesmo_horario.csv"))  # opcional
    pares_mesmo   = pares_mesmo_a + [p for p in pares_mesmo_b if p not in pares_mesmo_a]

    # Fixos
    fixos = carregar_fixos(os.path.join(dados_dir, "fixos.csv"))
    fixos = {d:b for d,b in fixos.items() if d in G}
    fora = {disc: b for disc, b in fixos.items() if b < 0 or b >= num_blocos}
    if fora:
        print(f"[AVISO] Fixos fora do calendário (0..{num_blocos-1}) ignorados: {fora}")
        fixos = {d:b for d,b in fixos.items() if 0 <= b < num_blocos}

    # Domínios por dia (dia_fixo.csv)
    dia_por_disc = carregar_dia_fixo(os.path.join(dados_dir, "dia_fixo.csv"))
    idx_dia = indice_blocos_por_dia(horarios)
    dominios = {}
    for disc, dia in dia_por_disc.items():
        if disc in G and dia in idx_dia:
            dominios[disc] = set(idx_dia[dia])

    # 3) Coloração / Alocação
    cores = colorir_grafo_balanceado(
        G,
        num_blocos=num_blocos,
        fixos=fixos,
        pares_mesmo_horario=pares_mesmo,   # tratado como “mesmo bloco”
        pares_mesmo_bloco=pares_mesmo,
        dominios_por_no=dominios,
        allow_extra_blocks=False,
        hard_fail=True
    )

    # 4) Saída — Disciplinas (console)
    print("\n--- Alocação das Disciplinas ---")
    for disc, bloco in sorted(cores.items(), key=lambda x: x[0]):
        label_disc = nome_exibicao.get(disc, disc)
        print(f"{label_disc} → {horarios.get(bloco, f'Bloco {bloco}')}")

    # 5) Estatísticas
    usados = sorted(set(cores.values()))
    dist = defaultdict(int)
    for _, b in cores.items():
        dist[b] += 1
    if usados:
        min_c = min(dist[b] for b in usados)
        max_c = max(dist[b] for b in usados)
        print(f"\nTotal de blocos usados: {len(usados)} (de {num_blocos})")
        print("Distribuição por bloco:", {b: dist[b] for b in usados})
        print("Desbalanceamento (max - min):", max_c - min_c)

    # 6) Exporta Excel/CSV com nome “alocacaohorario_{dias}x{blocos}”
    salvar_grade_excel_csv(
        cores,
        horarios,
        nome_exibicao,
        caminho_base=f"out/alocacaohorario_{dias_semana}x{blocos_dia}"
    )

if __name__ == "__main__":
    main()
