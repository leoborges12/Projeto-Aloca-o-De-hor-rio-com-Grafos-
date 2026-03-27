from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional, Literal, Dict, Any
from pathlib import Path
from datetime import datetime
from contextlib import redirect_stdout, redirect_stderr
from collections import defaultdict
import re
import csv
import sys
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from grafo import construir_grafo, colorir_grafo_balanceado
from main import montar_horarios, indice_blocos_por_dia

from supabase_client import supabase


# --------------------------
# Helpers
# --------------------------


class _Tee(io.StringIO):
    """
    Captura prints (stdout/stderr) e ao mesmo tempo continua escrevendo no terminal.
    """

    def __init__(self, real_stream):
        super().__init__()
        self.real_stream = real_stream

    def write(self, s):
        try:
            self.real_stream.write(s)
            self.real_stream.flush()
        except Exception:
            pass
        return super().write(s)


def _prof_display(prof_raw: str) -> str:
    toks = [t.strip() for t in re.split(r"[|,;/]+", str(prof_raw)) if t.strip()]
    uniq = list(dict.fromkeys(toks))
    return ", ".join(uniq)


def _norm_dia(dia: str) -> str:
    mapa = {
        "segunda": "segunda",
        "seg": "segunda",
        "terca": "terca",
        "terça": "terca",
        "ter": "terca",
        "quarta": "quarta",
        "qua": "quarta",
        "quinta": "quinta",
        "qui": "quinta",
        "sexta": "sexta",
        "sex": "sexta",
        "sabado": "sabado",
        "sábado": "sabado",
        "sab": "sabado",
        "domingo": "domingo",
        "dom": "domingo",
    }
    d = (dia or "").strip().lower()
    return mapa.get(d, "")


def _periodo_do_indice(indice_no_dia: int) -> str:
    return str(indice_no_dia + 1)


def _esc_csv(v):
    s = str(v or "")
    if '"' in s:
        s = s.replace('"', '""')
    if "," in s or "\n" in s or '"' in s:
        return f'"{s}"'
    return s


# --------------------------
# Modelos da API
# --------------------------


class Config(BaseModel):
    dias_semana: int = 5
    blocos_por_dia: int = 4
    conflito_por_prof: bool = True
    conflito_por_semestre: bool = True


class Disciplina(BaseModel):
    nome: str
    prof: str = ""
    semestre: str = ""
    aulas_por_semana: int = 1


class Restricao(BaseModel):
    tipo: Literal[
        "fixo", "dia_fixo", "nao_coincidir", "mesmo_bloco", "mesmo_horario"
    ] = "fixo"
    disciplina: Optional[str] = None
    bloco: Optional[int] = None
    ocorrencia: Optional[int] = None
    dia: Optional[str] = None
    disciplina1: Optional[str] = None
    disciplina2: Optional[str] = None


class Entrada(BaseModel):
    config: Config
    disciplinas: List[Disciplina]
    restricoes: List[Restricao] = []


class ImportarArquivosEntrada(BaseModel):
    arquivos: List[str]


class ExportarGradeEntrada(BaseModel):
    alocacao: Dict[str, int]
    horarios: Dict[str, str]
    nome_exibicao: Dict[str, str] = {}
    semestre_por_disc: Dict[str, str] = {}
    prefixo: str = "grade"


# --------------------------
# App / diretórios
# --------------------------

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
DADOS_DIR = BASE_DIR / "dados"
OUT_DIR = BASE_DIR / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# --------------------------
# Health / root
# --------------------------


@app.get("/")
def root():
    return {"ok": True, "docs": "/docs", "health": "/health"}


@app.get("/health")
def health():
    return {"ok": True}


@app.get("/supabase-test")
def supabase_test():
    try:
        res = supabase.table("cursos").select("*").limit(5).execute()
        return {"ok": True, "dados": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --------------------------
# CSV / datasets
# --------------------------


def _listar_datasets():
    if not DADOS_DIR.exists():
        return []
    nomes = set()
    for p in DADOS_DIR.glob("*_disciplinas.csv"):
        nome = p.name.replace("_disciplinas.csv", "")
        nomes.add(nome)
    return sorted(nomes)


def _ler_csv(path: Path):
    itens = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        amostra = f.read(2048)
        f.seek(0)

        try:
            dialeto = csv.Sniffer().sniff(amostra, delimiters=";,")
        except Exception:
            dialeto = csv.get_dialect("excel")

        reader = csv.DictReader(f, dialect=dialeto)
        for row in reader:
            limpo = {}
            for k, v in (row or {}).items():
                if k is None:
                    continue
                kk = str(k).strip()
                vv = v.strip() if isinstance(v, str) else v
                limpo[kk] = vv
            itens.append(limpo)

    return itens


def _ler_csv_texto(conteudo: str) -> list[dict]:
    itens = []
    f = io.StringIO(conteudo)

    amostra = conteudo[:2048]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=";,")
    except Exception:
        dialeto = csv.get_dialect("excel")

    reader = csv.DictReader(f, dialect=dialeto)

    for row in reader:
        limpo = {}
        for k, v in (row or {}).items():
            if k is None:
                continue
            kk = str(k).strip()
            vv = v.strip() if isinstance(v, str) else v
            limpo[kk] = vv
        itens.append(limpo)

    return itens


def _inferir_restricoes(rows: list[dict], nome_arquivo: str):
    fname = nome_arquivo.lower()
    tipo_padrao = None

    if "dia_fixo" in fname or ("dia" in fname and "fix" in fname):
        tipo_padrao = "dia_fixo"
    elif "fixo" in fname:
        tipo_padrao = "fixo"
    elif "mesmo_horario" in fname:
        tipo_padrao = "mesmo_horario"
    elif "mesmo_bloco" in fname:
        tipo_padrao = "mesmo_bloco"
    elif (
        "nao" in fname
        or "não" in fname
        or "coincidir" in fname
        or "restricoes" in fname
    ):
        tipo_padrao = "nao_coincidir"

    restricoes = []

    for row in rows:
        disciplina = row.get("disciplina") or row.get("Disciplina")
        disciplina1 = row.get("disciplina1") or row.get("Disciplina1")
        disciplina2 = row.get("disciplina2") or row.get("Disciplina2")
        bloco = row.get("bloco") or row.get("Bloco")
        dia = row.get("dia") or row.get("Dia")
        ocorrencia = (
            row.get("ocorrencia")
            or row.get("Ocorrencia")
            or row.get("ocorrência")
            or row.get("Ocorrência")
        )

        if isinstance(disciplina, str):
            disciplina = disciplina.strip() or None

        if isinstance(disciplina1, str):
            disciplina1 = disciplina1.strip() or None

        if isinstance(disciplina2, str):
            disciplina2 = disciplina2.strip() or None

        if isinstance(dia, str):
            dia = dia.strip() or None

        if isinstance(bloco, str):
            bloco = bloco.strip()
            if bloco == "":
                bloco = None
            else:
                bloco = int(bloco)

        if isinstance(ocorrencia, str):
            ocorrencia = ocorrencia.strip()
            if ocorrencia == "":
                ocorrencia = None
            else:
                ocorrencia = int(ocorrencia)

        r = {
            "tipo": tipo_padrao,
            "disciplina": disciplina,
            "disciplina1": disciplina1,
            "disciplina2": disciplina2,
            "bloco": bloco,
            "ocorrencia": ocorrencia,
            "dia": dia,
            "origem": nome_arquivo,
        }

        restricoes.append(r)
    return restricoes


def _normalizar_disciplina_row(r: dict) -> dict:
    nome_disc = (
        r.get("nome")
        or r.get("disciplina")
        or r.get("Nome")
        or r.get("Disciplina")
        or ""
    ).strip()

    prof = (
        r.get("prof") or r.get("professor") or r.get("Prof") or r.get("Professor") or ""
    ).strip()

    semestre = (r.get("semestre") or r.get("Semestre") or "").strip()

    aps = (
        r.get("aulas_por_semana")
        or r.get("aulasPorSemana")
        or r.get("ocorrencias_semanais")
        or r.get("quantidade_aulas")
        or 1
    )

    try:
        aps = max(1, int(aps))
    except Exception:
        aps = 1

    return {
        "nome": nome_disc,
        "prof": _prof_display(prof),
        "semestre": semestre,
        "aulas_por_semana": aps,
    }


# --------------------------
# Endpoints de datasets
# --------------------------


@app.get("/dados")
def listar_dados():
    return {"datasets": _listar_datasets()}


@app.get("/dados/{nome}")
def carregar_dados(nome: str):
    disc_path = DADOS_DIR / f"{nome}_disciplinas.csv"
    if not disc_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Dataset '{nome}' não encontrado (faltando {disc_path.name})",
        )

    disciplinas_raw = _ler_csv(disc_path)
    disciplinas = [_normalizar_disciplina_row(r) for r in disciplinas_raw]

    restricoes: list[dict] = []

    for p in sorted(DADOS_DIR.glob(f"{nome}_*.csv")):
        if p.name.endswith("_disciplinas.csv"):
            continue

        rows = _ler_csv(p)
        restricoes.extend(_inferir_restricoes(rows, p.name))

    return {"nome": nome, "disciplinas": disciplinas, "restricoes": restricoes}


# --------------------------
# Endpoints de restrições por arquivos
# --------------------------


def _listar_arquivos_restricoes_soltas():
    if not DADOS_DIR.exists():
        return []

    arquivos = []
    for p in sorted(DADOS_DIR.glob("*.csv")):
        if p.name.endswith("_disciplinas.csv"):
            continue
        arquivos.append(p.name)
    return arquivos


@app.get("/restricoes/arquivos")
def listar_arquivos_restricoes():
    return {"arquivos": _listar_arquivos_restricoes_soltas()}


@app.post("/restricoes/importar")
def importar_restricoes_por_arquivos(payload: ImportarArquivosEntrada):
    if not payload.arquivos:
        return {"restricoes": []}

    restricoes_total: list[dict] = []

    for nome_arq in payload.arquivos:
        p = (DADOS_DIR / nome_arq).resolve()
        if DADOS_DIR not in p.parents or not p.exists() or not p.is_file():
            raise HTTPException(
                status_code=404, detail=f"Arquivo não encontrado: {nome_arq}"
            )

        rows = _ler_csv(p)
        restrs = _inferir_restricoes(rows, p.name)

        for r in restrs:
            r["origem"] = p.name

        restricoes_total.extend(restrs)

    return {"restricoes": restricoes_total}


# --------------------------
# Upload de CSVs do computador
# --------------------------


@app.post("/upload/disciplinas")
async def upload_disciplinas(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Envie um arquivo .csv")

    raw = await file.read()
    texto = raw.decode("utf-8-sig", errors="replace")

    rows = _ler_csv_texto(texto)
    disciplinas = []

    for r in rows:
        d = _normalizar_disciplina_row(r)
        if d["nome"]:
            disciplinas.append(d)

    return {"disciplinas": disciplinas}


@app.post("/upload/restricoes")
async def upload_restricoes(files: List[UploadFile] = File(...)):
    if not files:
        return {"restricoes": []}

    restricoes: list[dict] = []

    for f in files:
        if not (f.filename or "").lower().endswith(".csv"):
            raise HTTPException(
                status_code=400, detail=f"Arquivo inválido (não é .csv): {f.filename}"
            )

        raw = await f.read()
        texto = raw.decode("utf-8-sig", errors="replace")
        rows = _ler_csv_texto(texto)

        restricoes.extend(_inferir_restricoes(rows, f.filename or "upload.csv"))

    return {"restricoes": restricoes}


# --------------------------
# Geração da grade
# --------------------------


@app.post("/gerar-grade")
def gerar_grade(dados: Entrada) -> Dict[str, Any]:
    tee_out = _Tee(sys.stdout)
    tee_err = _Tee(sys.stderr)

    try:
        with redirect_stdout(tee_out), redirect_stderr(tee_err):
            dias_semana = dados.config.dias_semana
            blocos_por_dia = dados.config.blocos_por_dia

            horarios = montar_horarios(dias_semana, blocos_por_dia)
            num_blocos = len(horarios)

            disciplinas_orig = [d.model_dump() for d in dados.disciplinas]

            disciplinas_list = []
            nome_base_por_expandida: Dict[str, str] = {}

            for d in disciplinas_orig:
                nome_base = d["nome"]
                prof = d.get("prof", "")
                semestre = d.get("semestre", "")
                aps = max(1, int(d.get("aulas_por_semana", 1) or 1))

                for i in range(aps):
                    nome_expandido = (
                        f"{nome_base} [{i+1}/{aps}]" if aps > 1 else nome_base
                    )

                    disciplinas_list.append(
                        {
                            "nome": nome_expandido,
                            "prof": prof,
                            "semestre": semestre,
                            "aulas_por_semana": 1,
                        }
                    )

                    nome_base_por_expandida[nome_expandido] = nome_base

            def expandir_nome_disciplina(nome: str) -> List[str]:
                if not nome:
                    return []
                encontrados = [
                    n for n, base in nome_base_por_expandida.items() if base == nome
                ]
                return encontrados if encontrados else [nome]

            def expandir_nome_disciplina_por_ocorrencia(
                nome: str, ocorrencia: Optional[int]
            ) -> List[str]:
                encontrados = expandir_nome_disciplina(nome)
                if not encontrados:
                    return []

                if ocorrencia is None:
                    return encontrados

                idx = int(ocorrencia) - 1
                if 0 <= idx < len(encontrados):
                    return [encontrados[idx]]

                return []

            G = construir_grafo(
                disciplinas_list,
                conflito_por_prof=dados.config.conflito_por_prof,
                conflito_por_semestre=dados.config.conflito_por_semestre,
            )

            fixos: Dict[str, int] = {}
            pares_mesmo: List[tuple] = []
            pares_nao: List[tuple] = []
            dia_por_disc: Dict[str, str] = {}

            for r in dados.restricoes:
                if r.tipo == "fixo":
                    if r.disciplina and r.bloco is not None:
                        expandidas = expandir_nome_disciplina_por_ocorrencia(
                            r.disciplina, r.ocorrencia
                        )
                        for nome_exp in expandidas:
                            fixos[nome_exp] = int(r.bloco)

                elif r.tipo == "dia_fixo":
                    if r.disciplina and r.dia:
                        dn = _norm_dia(r.dia)
                        if dn:
                            for nome_exp in expandir_nome_disciplina(r.disciplina):
                                dia_por_disc[nome_exp] = dn

                elif r.tipo == "nao_coincidir":
                    if r.disciplina1 and r.disciplina2:
                        a_list = expandir_nome_disciplina(r.disciplina1)
                        b_list = expandir_nome_disciplina(r.disciplina2)
                        for a in a_list:
                            for b in b_list:
                                if a != b:
                                    pares_nao.append((a, b))

                elif r.tipo in ("mesmo_bloco", "mesmo_horario"):
                    if r.disciplina1 and r.disciplina2:
                        a_list = expandir_nome_disciplina(r.disciplina1)
                        b_list = expandir_nome_disciplina(r.disciplina2)
                        for a in a_list:
                            for b in b_list:
                                if a != b:
                                    pares_mesmo.append((a, b))

            for a, b in pares_nao:
                if a in G and b in G:
                    G.add_edge(a, b)

            fixos = {d: b for d, b in fixos.items() if d in G}
            fixos = {d: b for d, b in fixos.items() if 0 <= b < num_blocos}

            dominios: Dict[str, set] = {}
            idx_dia = indice_blocos_por_dia(horarios)

            for disc, dia_norm in dia_por_disc.items():
                if disc in G and dia_norm in idx_dia:
                    dominios[disc] = set(idx_dia[dia_norm])

            cores = colorir_grafo_balanceado(
                G,
                num_blocos=num_blocos,
                fixos=fixos,
                pares_mesmo_horario=pares_mesmo,
                pares_mesmo_bloco=pares_mesmo,
                dominios_por_no=dominios,
                allow_extra_blocks=False,
                hard_fail=True,
            )

            dist = defaultdict(int)
            for _, b in cores.items():
                dist[b] += 1

            usados = sorted(dist.keys())
            desbalanceamento = 0
            if usados:
                desbalanceamento = max(dist[b] for b in usados) - min(
                    dist[b] for b in usados
                )

            total_disciplinas_base = len(disciplinas_orig)

            total_ocorrencias = sum(
                max(1, int(d.get("aulas_por_semana", 1) or 1)) for d in disciplinas_orig
            )

            ocorrencias_alocadas = len(
                [k for k in cores.keys() if k in nome_base_por_expandida]
            )

            disciplinas_base_alocadas = len(
                set(
                    nome_base_por_expandida.get(k, k)
                    for k in cores.keys()
                    if k in nome_base_por_expandida
                )
            )

            prof_display = {}
            for d in disciplinas_list:
                prof_display[d["nome"]] = _prof_display(d.get("prof", ""))

            nome_exibicao = {}
            for disc in G.nodes():
                nome_base = nome_base_por_expandida.get(disc, disc)
                prof_txt = prof_display.get(disc, "")
                nome_exibicao[disc] = (
                    f"{nome_base} / {prof_txt}" if prof_txt else nome_base
                )

            logs = (tee_out.getvalue() + "\n" + tee_err.getvalue()).strip()

            return {
                "alocacao": cores,
                "horarios": horarios,
                "stats": {
                    "total_blocos": num_blocos,
                    "blocos_usados": len(usados),
                    "desbalanceamento": desbalanceamento,
                    "dist_por_bloco": {str(k): int(v) for k, v in dist.items()},
                    "total_disciplinas_base": total_disciplinas_base,
                    "disciplinas_base_alocadas": disciplinas_base_alocadas,
                    "total_ocorrencias": total_ocorrencias,
                    "ocorrencias_alocadas": ocorrencias_alocadas,
                },
                "nome_exibicao": nome_exibicao,
                "logs": logs,
            }

    except Exception as e:
        logs = (tee_out.getvalue() + "\n" + tee_err.getvalue()).strip()
        msg = str(e)

        detail = ""
        if logs:
            detail += logs + "\n"
        detail += f"\nERRO: {msg}"

        raise HTTPException(status_code=400, detail=detail)


# --------------------------
# Exportação visual
# --------------------------


def _montar_grade_visual(
    alocacao: dict, horarios: dict, nome_exibicao: dict, semestre_por_disc: dict
):
    dias_ordem = ["seg", "ter", "qua", "qui", "sex", "sab", "dom"]

    dia_map = {
        "Seg": "seg",
        "Ter": "ter",
        "Qua": "qua",
        "Qui": "qui",
        "Sex": "sex",
        "Sab": "sab",
        "Dom": "dom",
    }

    grade = {}
    sem_semestre = {}

    def push_cell(container, semestre, periodo, dia, texto):
        if semestre not in container:
            container[semestre] = {}
        if periodo not in container[semestre]:
            container[semestre][periodo] = {}
        if dia not in container[semestre][periodo]:
            container[semestre][periodo][dia] = []
        container[semestre][periodo][dia].append(texto)

    def push_sem_semestre(container, periodo, dia, texto):
        if periodo not in container:
            container[periodo] = {}
        if dia not in container[periodo]:
            container[periodo][dia] = []
        container[periodo][dia].append(texto)

    total_blocos = len(horarios)
    blocos_por_dia = 4
    if total_blocos % 5 == 0 and total_blocos > 0:
        blocos_por_dia = total_blocos // 5

    for disc, bloco_raw in alocacao.items():
        bloco = int(bloco_raw)
        label = horarios.get(bloco, "")
        partes = str(label).split(" ")
        dia_txt = partes[0] if partes else ""
        dia = dia_map.get(dia_txt)
        if not dia:
            continue

        indice_no_dia = bloco % blocos_por_dia
        periodo = _periodo_do_indice(indice_no_dia)

        nome_base = re.sub(r"\s*\[\d+/\d+\]", "", disc)
        nome = nome_exibicao.get(disc, nome_base)

        semestre = str(semestre_por_disc.get(nome_base, "")).strip()

        if not semestre:
            push_sem_semestre(sem_semestre, periodo, dia, nome)
        else:
            push_cell(grade, semestre, periodo, dia, nome)

    return grade, sem_semestre, dias_ordem, blocos_por_dia


def _gerar_csv_visual(
    caminho_csv: Path,
    alocacao: dict,
    horarios: dict,
    nome_exibicao: dict,
    semestre_por_disc: dict,
):
    grade, sem_semestre, dias, blocos_por_dia = _montar_grade_visual(
        alocacao, horarios, nome_exibicao, semestre_por_disc
    )

    dias_uteis = dias[:5]
    semestres = sorted(grade.keys(), key=lambda x: str(x))
    periodos = [str(i + 1) for i in range(blocos_por_dia)]

    linhas = []

    linhas.append(_esc_csv("Oferta Regular — 2026 / 1"))
    linhas.append(",".join([_esc_csv(x) for x in ["Semestre", "Período", *dias_uteis]]))

    for sem in semestres:
        for p in periodos:
            row = [sem, p]
            for dia in dias_uteis:
                lista = grade.get(sem, {}).get(p, {}).get(dia, [])
                row.append("\n".join(lista))
            linhas.append(",".join(_esc_csv(x) for x in row))

    if sem_semestre:
        linhas.append("")
        linhas.append(_esc_csv("Disciplinas sem semestre informado"))
        linhas.append(",".join([_esc_csv(x) for x in ["Período", *dias_uteis]]))

        for p in periodos:
            row = [p]
            for dia in dias_uteis:
                lista = sem_semestre.get(p, {}).get(dia, [])
                row.append("\n".join(lista))
            linhas.append(",".join(_esc_csv(x) for x in row))

    caminho_csv.write_text("\n".join(linhas), encoding="utf-8-sig")


def _gerar_xlsx_visual(
    caminho_xlsx: Path,
    alocacao: dict,
    horarios: dict,
    nome_exibicao: dict,
    semestre_por_disc: dict,
):
    grade, sem_semestre, dias, blocos_por_dia = _montar_grade_visual(
        alocacao, horarios, nome_exibicao, semestre_por_disc
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Grade"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    dias_uteis = dias[:5]
    headers = ["Semestre", "Período", *dias_uteis]

    # título principal
    ws.merge_cells("A1:G1")
    ws["A1"] = "Oferta Regular — 2026 / 1"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # cabeçalho tabela principal
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    row_idx = 3
    semestres = sorted(grade.keys(), key=lambda x: str(x))
    periodos = [str(i + 1) for i in range(blocos_por_dia)]

    for sem in semestres:
        inicio_sem = row_idx

        for p in periodos:
            linha = [sem, p]
            for dia in dias_uteis:
                lista = grade.get(sem, {}).get(p, {}).get(dia, [])
                linha.append("\n".join(lista))
            ws.append(linha)

            for c in range(1, 8):
                cell = ws.cell(row=row_idx, column=c)
                cell.border = border
                cell.alignment = center if c <= 2 else top_left

            row_idx += 1

        if row_idx - 1 > inicio_sem:
            ws.merge_cells(
                start_row=inicio_sem, start_column=1, end_row=row_idx - 1, end_column=1
            )
            ws.cell(row=inicio_sem, column=1).alignment = center
            ws.cell(row=inicio_sem, column=1).font = bold
            ws.cell(row=inicio_sem, column=1).border = border

    # seção sem semestre informado
    if sem_semestre:
        row_idx += 2

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        ws.cell(row=row_idx, column=1, value="Disciplinas sem semestre informado")
        ws.cell(row=row_idx, column=1).font = title_font
        ws.cell(row=row_idx, column=1).alignment = center

        row_idx += 1
        headers2 = ["Período", *dias_uteis]
        for c, h in enumerate(headers2, start=1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.font = bold
            cell.alignment = center
            cell.border = border

        row_idx += 1
        for p in periodos:
            linha = [p]
            for dia in dias_uteis:
                lista = sem_semestre.get(p, {}).get(dia, [])
                linha.append("\n".join(lista))

            for c, valor in enumerate(linha, start=1):
                cell = ws.cell(row=row_idx, column=c, value=valor)
                cell.border = border
                cell.alignment = center if c == 1 else top_left

            row_idx += 1

    larguras = {
        "A": 12,
        "B": 12,
        "C": 32,
        "D": 32,
        "E": 32,
        "F": 32,
        "G": 32,
    }
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 34

    wb.save(caminho_xlsx)


@app.post("/exportar-grade")
def exportar_grade(payload: ExportarGradeEntrada) -> Dict[str, Any]:
    try:
        alocacao = {str(k): int(v) for k, v in payload.alocacao.items()}
        horarios = {int(k): str(v) for k, v in payload.horarios.items()}
        nome_exib = payload.nome_exibicao or {k: k for k in alocacao.keys()}
        semestre_por_disc = payload.semestre_por_disc or {}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Payload inválido: {e}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{payload.prefixo}_{ts}"

    caminho_csv = OUT_DIR / f"{base_name}.csv"
    caminho_xlsx = OUT_DIR / f"{base_name}.xlsx"

    try:
        _gerar_csv_visual(caminho_csv, alocacao, horarios, nome_exib, semestre_por_disc)
        _gerar_xlsx_visual(
            caminho_xlsx, alocacao, horarios, nome_exib, semestre_por_disc
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao salvar grade: {e}")

    return {
        "csv": f"/out/{caminho_csv.name}",
        "xlsx": f"/out/{caminho_xlsx.name}",
    }


@app.get("/out")
def listar_out():
    arquivos = sorted([p.name for p in OUT_DIR.glob("*") if p.is_file()])
    return {"arquivos": arquivos}


@app.get("/out/{arquivo}")
def baixar_out(arquivo: str):
    path = (OUT_DIR / arquivo).resolve()
    if OUT_DIR not in path.parents or not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    return FileResponse(path, filename=path.name)
